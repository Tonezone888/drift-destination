"""Build one working catalogue from the active destination workbook."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from zipfile import BadZipFile

from app_config import SHEET_NAMES, get_workbook_path

BUSINESS_SHEETS = {
    "Food & Drink Master": "Food & Drink",
    "Tours_Activities": "Tours & Activities",
    "Accommodation": "Accommodation",
    "Medical_Pharmacies": "Medical & Pharmacies",
    "Transport_Transfers": "Transport & Transfers",
    "Wellness_Massage_Spas": "Wellness, Massage & Spas",
    "Shops_Essentials": "Shops & Essentials",
    "Trades_Services": "Trades & Services",
    "Property_Lifestyle": "Property & Lifestyle",
    "Real_Estate": "Real Estate",
    "Real Estate": "Real Estate",
    "Community_Events": "Community & Events",
    "Weddings": "Weddings",
    "Places_of_Interest": "Places of Interest",
    "Inbox_Sort_Later": "Inbox (sort later)",
}



def normalize_food_type(type_val: str, sub_val: str = "") -> str:
    """Map Type to visitor tags: restaurant|cafe|bar|takeaway (comma-separated OK)."""
    raw = (sub_val or type_val or "").strip().lower()
    if not raw:
        return ""
    # already multi
    parts = [p.strip() for p in raw.replace("&", ",").replace("/", ",").split(",") if p.strip()]
    out = []
    for s in parts:
        if s in ("restaurant", "cafe", "bar", "takeaway"):
            if s not in out:
                out.append(s)
            continue
        if "takeaway" in s or "fast food" in s:
            if "takeaway" not in out:
                out.append("takeaway")
        if "cafe" in s or "bakery" in s or "coffee" in s or "dessert" in s:
            if "cafe" not in out:
                out.append("cafe")
        if "bar" in s or "pub" in s or "tavern" in s or "nightclub" in s:
            if "bar" not in out:
                out.append("bar")
        if "restaurant" in s or "dining" in s or "seafood" in s or "bistro" in s or "grill" in s:
            if "restaurant" not in out:
                out.append("restaurant")
        if "distillery" in s or "wine" in s:
            if "bar" not in out:
                out.append("bar")
    return ", ".join(out) if out else raw

    if "takeaway" in s or "fast food" in s or "fish and chip" in s or "kebab" in s:
        return "takeaway"
    if "cafe" in s or "bakery" in s or "coffee" in s or "dessert" in s or "ice cream" in s:
        return "cafe"
    if "bar" in s or "pub" in s or "tavern" in s or "nightclub" in s:
        return "bar"
    if "restaurant" in s or "dining" in s or "seafood" in s or "bistro" in s or "distillery" in s or "wine" in s:
        return "restaurant"
    return s


def normalize_poi_type(type_val: str, sub_val: str = "") -> str:
    """Places of Interest → visitor tokens: lookouts|trails|beaches|photo-spots|swimming."""
    s = (sub_val or type_val or "").strip().lower()
    if not s:
        return ""
    if "lookout" in s or "viewpoint" in s:
        return "lookouts"
    if "trail" in s or "track" in s or "walk" in s or "park" in s:
        return "trails"
    if "beach" in s:
        return "beaches"
    if "photo" in s or "sign" in s:
        return "photo-spots"
    if "swim" in s or "lagoon" in s or "hole" in s:
        return "swimming"
    if "meeting" in s:
        return "meeting-point"
    return s.replace(" ", "-")

NAME_COLUMNS = [
    "Business Name",
    "Venue",
    "Name",
    "Tour/Activity",
    "Business",
    "Property",
    "Operator",
]


def _clean(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _first(raw: dict, cols: list[str]) -> str:
    lower = {str(k).strip().lower(): k for k in raw}
    for c in cols:
        if c.lower() in lower:
            val = _clean(raw.get(lower[c.lower()]))
            if val:
                return val
    return ""


def build_catalogue(workbook_path: Path | str | None = None) -> tuple[pd.DataFrame, dict[str, Any]]:
    path = Path(workbook_path) if workbook_path else get_workbook_path()
    if not path.exists():
        return pd.DataFrame(), {
            "total_imported": 0,
            "sheet_counts": {},
            "error": "Workbook not found: %s" % path,
            "path": str(path),
        }

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except BadZipFile:
        return pd.DataFrame(), {
            "error": "Workbook file is damaged (not a valid Excel file). Restore the newest good file from Backups/ — see recovery steps.",
            "sheet_counts": {},
            "path": str(path),
        }
    except Exception as e:
        return pd.DataFrame(), {
            "error": "Could not open workbook: %s" % e,
            "sheet_counts": {},
            "path": str(path),
        }
    imported: list[dict[str, Any]] = []
    sheet_counts: dict[str, int] = {}
    skipped = 0

    try:
        # Travel map + any extra sheets (Clubs Members, Tradie Electrical, …)
        sheet_map = dict(BUSINESS_SHEETS)
        for sn in wb.sheetnames:
            if sn not in sheet_map and not str(sn).startswith("_"):
                # Human label from sheet title
                sheet_map[sn] = str(sn).replace("_", " ")
        for sheet_name, category in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                sheet_counts[sheet_name] = 0
                continue
            rows = wb[sheet_name].iter_rows(values_only=True)
            try:
                headings = [_clean(h) for h in next(rows)]
            except StopIteration:
                sheet_counts[sheet_name] = 0
                continue
            count = 0
            for excel_row, values in enumerate(rows, start=2):
                raw = {
                    headings[i]: values[i] if i < len(values) else ""
                    for i in range(len(headings))
                    if headings[i]
                }
                name = _first(raw, NAME_COLUMNS)
                if not name:
                    skipped += 1
                    continue
                phone = _first(raw, ["Phone", "Phone Number", "Mobile"])
                website = _first(raw, ["Official Website", "Website", "Property Website"])
                email = _first(raw, ["Email"])
                locality = _first(raw, ["Locality", "Location"])
                address = _first(raw, ["Street Address", "Address", "Address / Base"])
                status = _first(raw, ["Status", "Production Status"]) or "Active"
                notes = _first(raw, ["Notes", "Description / Notes", "Description"])
                lat = _first(raw, ["Latitude"])
                lng = _first(raw, ["Longitude"])
                hours = _first(raw, ["Opening Hours"])
                fb = _first(raw, ["Facebook"])
                ig = _first(raw, ["Instagram"])
                place_id = _first(raw, ["Google Place ID"])
                photo_name = _first(raw, ["Photo Name", "Photo Ref", "Photo Name (ref only)"])
                rating = _first(raw, ["Google Rating", "Rating"])
                review_count = _first(raw, ["Google Review Count", "Review Count", "Reviews"])
                maps_url = _first(raw, ["Google Maps URL", "Maps URL"])
                verified = _first(raw, ["Verified", "Subscribed", "Is Verified", "Subscriber"])
                last_pay = _first(raw, ["Last Payment", "Last Payment At", "Subscription Paid", "Paid At"])
                sub_until = _first(raw, ["Subscription Until", "Sub Until", "Verified Until"])
                btype = _first(raw, ["Type", "Primary category"])
                subcat = _first(raw, ["Subcategory", "Sub category", "Drift Subcategory"])
                if category == "Food & Drink":
                    subcat = normalize_food_type(btype, subcat)
                elif category == "Places of Interest":
                    subcat = normalize_poi_type(btype, subcat)
                elif not subcat:
                    subcat = btype
                imported.append(
                    {
                        "Drift ID": "%s:%s" % (sheet_name, excel_row),
                        "Business Name": name,
                        "Category": category,
                        "Locality": locality,
                        "Address": address,
                        "Phone": phone,
                        "Website": website,
                        "Email": email,
                        "Status": status,
                        "Notes": notes,
                        "Latitude": lat,
                        "Longitude": lng,
                        "Opening Hours": hours,
                        "Facebook": fb,
                        "Instagram": ig,
                        "Google Place ID": place_id,
                        "Photo Name": photo_name,
                        "Google Rating": rating,
                        "Google Review Count": review_count,
                        "Google Maps URL": maps_url,
                        "Verified": verified,
                        "Last Payment": last_pay,
                        "Subscription Until": sub_until,
                        "Type": btype,
                        "Subcategory": subcat,
                        "Source Sheet": sheet_name,
                        "Source Row": excel_row,
                    }
                )
                count += 1
            sheet_counts[sheet_name] = count
    finally:
        wb.close()

    df = pd.DataFrame(imported)
    return df, {
        "total_imported": len(df),
        "sheet_counts": sheet_counts,
        "skipped_blank_rows": skipped,
        "path": str(path),
    }
